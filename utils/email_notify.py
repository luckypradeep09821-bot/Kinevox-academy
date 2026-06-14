"""
utils/email_notify.py
======================
Email notification helper for Kinevox Academy.

Sends an email (via the Brevo transactional email HTTP API) with the new
record data as an Excel (.xlsx) attachment whenever a new Employee or
Student is created. Brevo is used instead of raw SMTP because most cloud
hosts (e.g. Render free tier) block outbound SMTP ports 25/465/587 —
Brevo's API runs over plain HTTPS (port 443) which is never blocked.

Configuration (via .env):
  BREVO_API_KEY         — Brevo API key                 (required)
  NOTIFY_EMAIL_TO       — recipient address              (required)
  NOTIFY_FROM_EMAIL     — verified sender address in Brevo (required)
  NOTIFY_FROM_NAME      — Friendly sender name (default: Kinevox Academy)

Usage:
  from utils.email_notify import notify_new_record
  notify_new_record("employee", emp_dict)   # fire-and-forget (threaded)
  notify_new_record("student",  stu_dict)
"""

import base64
import io
import os
import threading
from datetime import datetime

import requests

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False


BREVO_API_URL = "https://api.brevo.com/v3/smtp/email"


# ── Configuration helpers ─────────────────────────────────────────────────────

def _cfg(key: str, default: str = "") -> str:
    return os.getenv(key, default).strip()


# ── Brevo API sender ───────────────────────────────────────────────────────────

def _send_via_brevo(
    to_addr: str,
    subject: str,
    plain_body: str,
    html_body: str,
    attachments: list | None = None,
) -> tuple[bool, str]:
    """
    Send an email via the Brevo HTTP API (port 443 — never blocked by hosts).

    attachments: list of (filename, bytes) tuples. Optional.
    Returns (success, message).
    """
    api_key   = _cfg("BREVO_API_KEY")
    from_addr = _cfg("NOTIFY_FROM_EMAIL")
    from_name = _cfg("NOTIFY_FROM_NAME", "Kinevox Academy")

    if not all([api_key, to_addr, from_addr]):
        return False, "BREVO_API_KEY / NOTIFY_EMAIL_TO / NOTIFY_FROM_EMAIL not set — skipping."

    payload = {
        "sender": {"name": from_name, "email": from_addr},
        "to": [{"email": to_addr}],
        "subject": subject,
        "htmlContent": html_body,
        "textContent": plain_body,
    }

    if attachments:
        payload["attachment"] = [
            {
                "name": filename,
                "content": base64.b64encode(file_bytes).decode("ascii"),
            }
            for filename, file_bytes in attachments
        ]

    try:
        resp = requests.post(
            BREVO_API_URL,
            json=payload,
            headers={
                "api-key": api_key,
                "Content-Type": "application/json",
                "Accept": "application/json",
            },
            timeout=30,
        )
        if resp.status_code in (200, 201, 202):
            return True, "sent"
        return False, f"Brevo API error {resp.status_code}: {resp.text[:300]}"
    except Exception as exc:
        return False, str(exc)


# ── Excel builder ─────────────────────────────────────────────────────────────

def _build_excel(record_type: str, data: dict) -> bytes:
    """Return an .xlsx file as bytes containing the record fields."""
    if not _HAS_OPENPYXL:
        raise RuntimeError("openpyxl is not installed — cannot build Excel attachment")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = record_type.capitalize()

    # ── Styles ────────────────────────────────────────────────────────────────
    header_fill  = PatternFill("solid", fgColor="6C63FF")
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    label_font   = Font(bold=True, color="333333", size=10)
    value_font   = Font(color="111111", size=10)
    center       = Alignment(horizontal="center", vertical="center")
    left         = Alignment(horizontal="left",   vertical="center")
    thin         = Side(style="thin", color="CCCCCC")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Title row ─────────────────────────────────────────────────────────────
    title_text = (
        f"New {'Employee' if record_type == 'employee' else 'Student'} Record"
        f" — Kinevox Academy"
    )
    ws.merge_cells("A1:B1")
    title_cell = ws["A1"]
    title_cell.value = title_text
    title_cell.font  = Font(bold=True, color="FFFFFF", size=13)
    title_cell.fill  = header_fill
    title_cell.alignment = center
    ws.row_dimensions[1].height = 28

    # ── Timestamp row ─────────────────────────────────────────────────────────
    ws.merge_cells("A2:B2")
    ts_cell = ws["A2"]
    ts_cell.value = f"Generated: {datetime.now().strftime('%d %b %Y, %I:%M %p')}"
    ts_cell.font  = Font(italic=True, color="888888", size=9)
    ts_cell.alignment = center
    ws.row_dimensions[2].height = 18

    # Spacer
    ws.row_dimensions[3].height = 8

    # ── Column headers ────────────────────────────────────────────────────────
    ws["A4"].value = "Field"
    ws["B4"].value = "Value"
    for col in ("A4", "B4"):
        cell = ws[col]
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = border
    ws.row_dimensions[4].height = 22

    # ── Field rows ────────────────────────────────────────────────────────────
    SKIP = {"photo", "sign", "idDoc", "resume", "courses"}

    # Human-friendly field name mapping
    LABELS = {
        # Employee
        "id":          "Employee ID",
        "name":        "Full Name",
        "role":        "Role / Designation",
        "dept":        "Department",
        "email":       "Email Address",
        "phone":       "Phone Number",
        "doj":         "Date of Joining",
        "salary":      "Basic Salary (₹)",
        "bonus":       "Bonus (₹)",
        "credits":     "Credits",
        "schedule":    "Schedule",
        "status":      "Status",
        # Student
        "institution": "Institution",
        "course":      "Course",
        "assigned_to": "Trainer ID",
        "trainer":     "Trainer Name",
        "trainer_name":"Trainer Name",
        "enroll_date": "Enroll Date",
        "progress":    "Progress (%)",
        "percentage":  "Completion (%)",
        "dob":         "Date of Birth",
    }

    row = 5
    for key, value in data.items():
        if key in SKIP:
            continue
        if value is None or value == "":
            continue

        label = LABELS.get(key, key.replace("_", " ").title())
        val   = str(value)

        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font      = label_font
        label_cell.alignment = left
        label_cell.border    = border

        value_cell = ws.cell(row=row, column=2, value=val)
        value_cell.font      = value_font
        value_cell.alignment = left
        value_cell.border    = border

        ws.row_dimensions[row].height = 20
        row += 1

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 42

    # ── Save to bytes ─────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Email sender ──────────────────────────────────────────────────────────────

def _send_email(record_type: str, data: dict) -> None:
    """Build and send the notification email (blocking — run in a thread)."""
    to_addr   = _cfg("NOTIFY_EMAIL_TO")
    from_addr = _cfg("NOTIFY_FROM_EMAIL")
    api_key   = _cfg("BREVO_API_KEY")

    if not all([to_addr, from_addr, api_key]):
        print("[email_notify] BREVO_API_KEY / NOTIFY_EMAIL_TO / NOTIFY_FROM_EMAIL not set — skipping.")
        return

    # ── Build email ───────────────────────────────────────────────────────────
    label      = "Employee" if record_type == "employee" else "Student"
    rec_name   = data.get("name", "Unknown")
    rec_id     = data.get("id", "—")
    subject    = f"[Kinevox Academy] New {label} Added: {rec_name} ({rec_id})"
    timestamp  = datetime.now().strftime("%d %b %Y at %I:%M %p")

    # Plain-text fallback
    plain_lines = [f"New {label} record added on {timestamp}.", ""]
    for k, v in data.items():
        if k in {"photo", "sign", "idDoc", "resume", "courses"} or v is None or v == "":
            continue
        plain_lines.append(f"  {k}: {v}")
    plain_lines += ["", "—", "Kinevox Academy Management System"]
    plain_body = "\n".join(plain_lines)

    # HTML body
    rows_html = ""
    SKIP = {"photo", "sign", "idDoc", "resume", "courses"}
    LABELS = {
        "id":"ID","name":"Full Name","role":"Role","dept":"Department",
        "email":"Email","phone":"Phone","doj":"Date of Joining",
        "salary":"Basic Salary (₹)","bonus":"Bonus (₹)","credits":"Credits",
        "schedule":"Schedule","status":"Status","institution":"Institution",
        "course":"Course","assigned_to":"Trainer ID","trainer":"Trainer",
        "trainer_name":"Trainer Name","enroll_date":"Enroll Date",
        "progress":"Progress (%)","percentage":"Completion (%)","dob":"Date of Birth",
    }
    for k, v in data.items():
        if k in SKIP or v is None or v == "":
            continue
        friendly = LABELS.get(k, k.replace("_", " ").title())
        rows_html += (
            f'<tr><td style="padding:9px 14px;font-weight:600;color:#444;'
            f'background:#f8f8fc;border-bottom:1px solid #eee;white-space:nowrap">'
            f'{friendly}</td>'
            f'<td style="padding:9px 14px;color:#222;border-bottom:1px solid #eee">'
            f'{v}</td></tr>'
        )

    html_body = f"""
<!DOCTYPE html>
<html>
<body style="margin:0;padding:0;background:#f0f0f7;font-family:'DM Sans',Arial,sans-serif">
<table width="100%" cellpadding="0" cellspacing="0">
<tr><td align="center" style="padding:32px 16px">
<table width="560" cellpadding="0" cellspacing="0"
  style="background:#fff;border-radius:12px;overflow:hidden;
         box-shadow:0 4px 24px rgba(0,0,0,.08)">

  <!-- Header -->
  <tr>
    <td style="background:linear-gradient(135deg,#6c63ff,#38bdf8);
               padding:28px 32px;text-align:center">
      <div style="font-size:22px;font-weight:800;color:#fff;letter-spacing:-.5px">
        🎓 Kinevox Academy
      </div>
      <div style="color:rgba(255,255,255,.85);font-size:13px;margin-top:4px">
        Management System Notification
      </div>
    </td>
  </tr>

  <!-- Alert banner -->
  <tr>
    <td style="background:#f0edff;padding:14px 32px;border-bottom:2px solid #6c63ff">
      <span style="font-size:15px;font-weight:700;color:#6c63ff">
        ✅ New {label} Record Added
      </span>
      <span style="font-size:12px;color:#888;float:right;margin-top:2px">
        {timestamp}
      </span>
    </td>
  </tr>

  <!-- Data table -->
  <tr>
    <td style="padding:24px 32px">
      <table width="100%" cellpadding="0" cellspacing="0"
             style="border:1px solid #e8e8f0;border-radius:8px;overflow:hidden">
        {rows_html}
      </table>
    </td>
  </tr>

  <!-- Footer -->
  <tr>
    <td style="background:#f8f8fc;padding:16px 32px;text-align:center;
               font-size:11px;color:#aaa;border-top:1px solid #eee">
      The full record is attached as an Excel file for your records.<br>
      This is an automated message from Kinevox Academy Management System.
    </td>
  </tr>

</table>
</td></tr>
</table>
</body>
</html>
"""

    # ── Build attachment (Excel) ─────────────────────────────────────────────
    attachments = []
    try:
        xlsx_bytes = _build_excel(record_type, data)
        filename   = f"kinevox_{record_type}_{rec_id}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        attachments.append((filename, xlsx_bytes))
    except Exception as exc:
        print(f"[email_notify] Excel build failed: {exc} — sending without attachment.")

    # ── Send via Brevo ────────────────────────────────────────────────────────
    ok, info = _send_via_brevo(to_addr, subject, plain_body, html_body, attachments)
    if ok:
        print(f"[email_notify] ✅ Notification sent to {to_addr} for {record_type} {rec_id}")
    else:
        print(f"[email_notify] ❌ Failed to send email: {info}")


# ── Public API ────────────────────────────────────────────────────────────────

def notify_new_record(record_type: str, data: dict) -> None:
    """
    Fire-and-forget: sends an email notification in a background thread.

    Args:
        record_type: "employee" or "student"
        data:        The record dict (same as what was saved to DB)
    """
    t = threading.Thread(
        target=_send_email,
        args=(record_type, data),
        name=f"email-notify-{record_type}",
    )
    t.daemon = False
    t.start()


# ── Document upload notification ───────────────────────────────────────────────

def _send_document_email(
    record_type: str,
    record: dict,
    doc_type_label: str,
    file_bytes: bytes,
    original_filename: str,
) -> None:
    """Send an email with the uploaded document as an attachment."""
    to_addr   = _cfg("NOTIFY_EMAIL_TO")
    from_addr = _cfg("NOTIFY_FROM_EMAIL")
    api_key   = _cfg("BREVO_API_KEY")

    if not all([to_addr, from_addr, api_key]):
        print("[email_notify] BREVO_API_KEY / NOTIFY_EMAIL_TO / NOTIFY_FROM_EMAIL not set — skipping.")
        return

    label    = "Employee" if record_type == "employee" else "Student"
    rec_name = record.get("name", "Unknown")
    rec_id   = record.get("id", "—")
    dept     = record.get("dept") or record.get("course") or "—"
    role     = record.get("role") or record.get("institution") or "—"
    timestamp = datetime.now().strftime("%d %b %Y at %I:%M %p")

    subject = (
        f"[Kinevox Academy] Document Upload: {doc_type_label} "
        f"— {rec_name} ({rec_id})"
    )

    plain_body = (
        f"A new document was uploaded for {label}: {rec_name} ({rec_id})\n"
        f"Document Type : {doc_type_label}\n"
        f"File          : {original_filename}\n"
        f"Uploaded at   : {timestamp}\n\n"
        f"The document is attached to this email.\n"
        f"— Kinevox Academy Management System"
    )

    html_body = f"""<!DOCTYPE html>
<html>
<body style="margin:0;padding:0;background:#f0f0f7;font-family:'DM Sans',Arial,sans-serif">
<table width="100%" cellpadding="0" cellspacing="0">
<tr><td align="center" style="padding:32px 16px">
<table width="560" cellpadding="0" cellspacing="0"
  style="background:#fff;border-radius:12px;overflow:hidden;
         box-shadow:0 4px 24px rgba(0,0,0,.08)">
  <tr>
    <td style="background:linear-gradient(135deg,#6c63ff,#38bdf8);padding:28px 32px;text-align:center">
      <div style="font-size:22px;font-weight:800;color:#fff">🎓 Kinevox Academy</div>
      <div style="color:rgba(255,255,255,.85);font-size:13px;margin-top:4px">Document Upload Notification</div>
    </td>
  </tr>
  <tr>
    <td style="background:#f0edff;padding:14px 32px;border-bottom:2px solid #6c63ff">
      <span style="font-size:15px;font-weight:700;color:#6c63ff">📎 New Document Uploaded</span>
      <span style="font-size:12px;color:#888;float:right;margin-top:2px">{timestamp}</span>
    </td>
  </tr>
  <tr>
    <td style="padding:24px 32px">
      <table width="100%" cellpadding="0" cellspacing="0"
             style="border:1px solid #e8e8f0;border-radius:8px;overflow:hidden">
        <tr>
          <td style="padding:10px 14px;font-weight:600;color:#444;background:#f8f8fc;border-bottom:1px solid #eee;white-space:nowrap">{label} Name</td>
          <td style="padding:10px 14px;color:#222;border-bottom:1px solid #eee">{rec_name}</td>
        </tr>
        <tr>
          <td style="padding:10px 14px;font-weight:600;color:#444;background:#f8f8fc;border-bottom:1px solid #eee;white-space:nowrap">{label} ID</td>
          <td style="padding:10px 14px;color:#222;border-bottom:1px solid #eee">{rec_id}</td>
        </tr>
        <tr>
          <td style="padding:10px 14px;font-weight:600;color:#444;background:#f8f8fc;border-bottom:1px solid #eee;white-space:nowrap">{'Department' if record_type=='employee' else 'Institution'}</td>
          <td style="padding:10px 14px;color:#222;border-bottom:1px solid #eee">{dept}</td>
        </tr>
        <tr>
          <td style="padding:10px 14px;font-weight:600;color:#444;background:#f8f8fc;border-bottom:1px solid #eee;white-space:nowrap">Document Type</td>
          <td style="padding:10px 14px;color:#222;border-bottom:1px solid #eee">{doc_type_label}</td>
        </tr>
        <tr>
          <td style="padding:10px 14px;font-weight:600;color:#444;background:#f8f8fc;white-space:nowrap">File Name</td>
          <td style="padding:10px 14px;color:#222">{original_filename}</td>
        </tr>
      </table>
      <div style="margin-top:16px;padding:14px 18px;background:#f0fff4;border:1px solid #6ee7b7;
                  border-radius:8px;font-size:13px;color:#065f46">
        📎 The uploaded document is attached to this email.
      </div>
    </td>
  </tr>
  <tr>
    <td style="background:#f8f8fc;padding:16px 32px;text-align:center;
               font-size:11px;color:#aaa;border-top:1px solid #eee">
      This is an automated message from Kinevox Academy Management System.
    </td>
  </tr>
</table>
</td></tr>
</table>
</body>
</html>"""

    # ── Send via Brevo (file passed in-memory — no disk read needed) ──────────
    ok, info = _send_via_brevo(
        to_addr, subject, plain_body, html_body,
        attachments=[(original_filename, file_bytes)],
    )
    if ok:
        print(f"[email_notify] ✅ Document email sent for {record_type} {rec_id} — {doc_type_label}")
    else:
        print(f"[email_notify] ❌ Failed to send document email: {info}")


def notify_document_upload(
    record_type: str,
    record: dict,
    doc_type_label: str,
    file_bytes: bytes,
    original_filename: str,
) -> None:
    """Fire-and-forget document upload email notification."""
    t = threading.Thread(
        target=_send_document_email,
        args=(record_type, record, doc_type_label, file_bytes, original_filename),
        name="email-notify-doc",
    )
    t.daemon = False
    t.start()


# ── Batch document upload notification ────────────────────────────────────────

def _send_documents_batch_email(
    record_type: str,
    record: dict,
    documents: list,   # list of (doc_type_label, file_bytes, original_filename)
) -> None:
    """Send one email with ALL uploaded documents as attachments."""
    to_addr   = _cfg("NOTIFY_EMAIL_TO")
    from_addr = _cfg("NOTIFY_FROM_EMAIL")
    api_key   = _cfg("BREVO_API_KEY")

    if not all([to_addr, from_addr, api_key]):
        print("[email_notify] BREVO_API_KEY / NOTIFY_EMAIL_TO / NOTIFY_FROM_EMAIL not set — skipping batch doc email.")
        return

    label     = "Employee" if record_type == "employee" else "Student"
    rec_name  = record.get("name", "Unknown")
    rec_id    = record.get("id", "—")
    dept      = record.get("dept") or record.get("course") or "—"
    timestamp = datetime.now().strftime("%d %b %Y at %I:%M %p")
    doc_count = len(documents)

    subject = (
        f"[Kinevox Academy] {doc_count} Document{'s' if doc_count != 1 else ''} Uploaded"
        f" — {rec_name} ({rec_id})"
    )

    # Build rows for each document
    doc_rows_html = ""
    doc_lines_plain = []
    for i, (dtype_label, file_bytes, orig_name) in enumerate(documents, 1):
        bg = "#f8f8fc" if i % 2 == 1 else "#fff"
        doc_rows_html += f"""
        <tr>
          <td style="padding:9px 14px;font-weight:600;color:#444;background:{bg};
                     border-bottom:1px solid #eee;white-space:nowrap">
            📎 {dtype_label}
          </td>
          <td style="padding:9px 14px;color:#222;border-bottom:1px solid #eee">
            {orig_name}
          </td>
        </tr>"""
        doc_lines_plain.append(f"  [{i}] {dtype_label}: {orig_name}")

    plain_body = (
        f"{doc_count} document(s) uploaded for {label}: {rec_name} ({rec_id})\n"
        f"Uploaded at: {timestamp}\n\n"
        + "\n".join(doc_lines_plain)
        + "\n\nAll files are attached to this email.\n"
        "— Kinevox Academy Management System"
    )

    html_body = f"""<!DOCTYPE html>
<html>
<body style="margin:0;padding:0;background:#f0f0f7;font-family:'DM Sans',Arial,sans-serif">
<table width="100%" cellpadding="0" cellspacing="0">
<tr><td align="center" style="padding:32px 16px">
<table width="560" cellpadding="0" cellspacing="0"
  style="background:#fff;border-radius:12px;overflow:hidden;
         box-shadow:0 4px 24px rgba(0,0,0,.08)">
  <tr>
    <td style="background:linear-gradient(135deg,#6c63ff,#38bdf8);padding:28px 32px;text-align:center">
      <div style="font-size:22px;font-weight:800;color:#fff">🎓 Kinevox Academy</div>
      <div style="color:rgba(255,255,255,.85);font-size:13px;margin-top:4px">Document Upload Notification</div>
    </td>
  </tr>
  <tr>
    <td style="background:#f0edff;padding:14px 32px;border-bottom:2px solid #6c63ff">
      <span style="font-size:15px;font-weight:700;color:#6c63ff">
        📎 {doc_count} Document{'s' if doc_count != 1 else ''} Uploaded
      </span>
      <span style="font-size:12px;color:#888;float:right;margin-top:2px">{timestamp}</span>
    </td>
  </tr>
  <tr>
    <td style="padding:24px 32px">
      <!-- Person info -->
      <table width="100%" cellpadding="0" cellspacing="0"
             style="border:1px solid #e8e8f0;border-radius:8px;overflow:hidden;margin-bottom:20px">
        <tr>
          <td style="padding:9px 14px;font-weight:600;color:#444;background:#f8f8fc;
                     border-bottom:1px solid #eee;white-space:nowrap">{label} Name</td>
          <td style="padding:9px 14px;color:#222;border-bottom:1px solid #eee">{rec_name}</td>
        </tr>
        <tr>
          <td style="padding:9px 14px;font-weight:600;color:#444;background:#f8f8fc;
                     border-bottom:1px solid #eee;white-space:nowrap">{label} ID</td>
          <td style="padding:9px 14px;color:#222;border-bottom:1px solid #eee">{rec_id}</td>
        </tr>
        <tr>
          <td style="padding:9px 14px;font-weight:600;color:#444;background:#f8f8fc;
                     white-space:nowrap">{'Department' if record_type == 'employee' else 'Institution'}</td>
          <td style="padding:9px 14px;color:#222">{dept}</td>
        </tr>
      </table>

      <!-- Documents list -->
      <div style="font-size:12px;font-weight:700;color:#6c63ff;text-transform:uppercase;
                  letter-spacing:1px;margin-bottom:8px">Uploaded Documents</div>
      <table width="100%" cellpadding="0" cellspacing="0"
             style="border:1px solid #e8e8f0;border-radius:8px;overflow:hidden">
        {doc_rows_html}
      </table>

      <div style="margin-top:16px;padding:14px 18px;background:#f0fff4;
                  border:1px solid #6ee7b7;border-radius:8px;font-size:13px;color:#065f46">
        📎 All {doc_count} file{'s are' if doc_count != 1 else ' is'} attached to this email.
      </div>
    </td>
  </tr>
  <tr>
    <td style="background:#f8f8fc;padding:16px 32px;text-align:center;
               font-size:11px;color:#aaa;border-top:1px solid #eee">
      This is an automated message from Kinevox Academy Management System.
    </td>
  </tr>
</table>
</td></tr>
</table>
</body>
</html>"""

    # ── Send via Brevo (all files passed in-memory) ────────────────────────────
    attachments = [(orig_name, file_bytes) for _, file_bytes, orig_name in documents]
    ok, info = _send_via_brevo(to_addr, subject, plain_body, html_body, attachments)
    if ok:
        print(f"[email_notify] ✅ Batch doc email sent — {doc_count} file(s) for {record_type} {rec_id}")
    else:
        print(f"[email_notify] ❌ Batch doc email failed: {info}")


def notify_documents_batch(
    record_type: str,
    record: dict,
    documents: list,
) -> None:
    """Fire-and-forget: send one email with all uploaded documents attached."""
    t = threading.Thread(
        target=_send_documents_batch_email,
        args=(record_type, record, documents),
        name="email-notify-docs-batch",
    )
    t.daemon = False
    t.start()
